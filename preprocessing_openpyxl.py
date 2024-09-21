import os
import openpyxl
from pathlib import Path

from logger import logger
import config
from utility_functions import terminate_script, catch_errors


@catch_errors()
def preprocessing_file_excel(path_file_excel: str):
    file_excel =Path(path_file_excel).name
    workbook = None
    try:
        workbook = openpyxl.load_workbook(path_file_excel)
    except Exception as e:
        terminate_script(f'''{file_excel}: Ошибка обработки файла. Возможно открыт обрабатываемый файл. Закройте этот файл и снова запустите скрипт.
                              Ошибка: {e}''')

    sheet = workbook.active

    # Снимаем объединение ячеек
    merged_cells_ranges = list(sheet.merged_cells.ranges)
    for merged_cell_range in merged_cells_ranges:
        sheet.unmerge_cells(str(merged_cell_range))

    # Столбец с уровнями группировок
    sheet.insert_cols(idx=1)
    for row_index in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row_index, column=1)
        cell.value = sheet.row_dimensions[row_index].outline_level
    sheet['A1'] = "Уровень"

    # Столбец с признаком курсив
    sheet.insert_cols(idx=2)
    sheet['B1'] = "Курсив"

    kor_schet_col_index =None

    for row in sheet.iter_rows(values_only=True):
        if "Счет" in row:
            try:
                kor_schet_col_index = row.index("Кор.счет") + 1  # We add 1 because indexing starts from 0
            except ValueError:
                try:
                    kor_schet_col_index = row.index("Кор. Счет") + 1
                except ValueError:
                    terminate_script(f'''{file_excel}: Не обнаружен столбец 'Кор.счет' или 'Кор. Счет'.
                    Убедитесь, что обрабатываемый файл является Анализом счета.
                    Иначе сообщите разработчику,
                    если имя столбца с корреспондирующими счетами отличается от Кор.счет или Кор. Счет''')
            break
    else:

        terminate_script(f'''{file_excel}: Не обнаружена строка в файле, содержащая поле 'Счет'.
                            Убедитесь, что обрабатываемый файл является Анализом счета.
                            Также возможно, что в анализе счета не указана группировка. 
                            Сообщите разработчику,
                            если имя столбца отличается от Счет''')

    # Мы заполняем новый столбец значениями, основанными на форматировании ячеек курсивом
    for row_index in range(2, sheet.max_row + 1):  # We start with 2 to skip the title
        kor_schet_cell = sheet.cell(row=row_index, column=kor_schet_col_index)
        new_cell = sheet.cell(row=row_index, column=2)
        new_cell.value = 1 if kor_schet_cell.font and kor_schet_cell.font.italic else 0

    file_excel_treatment = f'{config.folder_path_preprocessing}/preprocessing_{file_excel}'
    if not os.path.exists(f'{config.folder_path_preprocessing}'):
        os.makedirs(f'{config.folder_path_preprocessing}')
    workbook.save(file_excel_treatment)
    workbook.close()
    logger.info(f'{file_excel}: сняли объединение ячеек, проставили уровни группировок и признак курсив в ячейках')
    return file_excel_treatment
